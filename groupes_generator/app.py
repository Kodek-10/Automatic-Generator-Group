import os
import random
from flask import Flask, render_template, request, redirect, url_for, flash, send_from_directory
from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

app = Flask(__name__)
app.secret_key = 'votre_cle_secrete'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['ALLOWED_EXTENSIONS'] = {'xlsx', 'xls'}

# Créer le dossier uploads s'il n'existe pas
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Variables globales pour stocker les données entre les étapes
etudiants = []
groupes = []
parametres = {}

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def repartition_aleatoire(etudiants, taille_groupe, min_garcons):
    # Initialisation de la variable groupes
    groupes = []
    
    # Vérification préalable du format des données
    if not etudiants or not all(isinstance(e, tuple) and len(e) == 3 for e in etudiants):
        print("Format des données étudiants invalide!")
        return groupes  # Retourne une liste vide

    # Séparation garçons/filles
    garcons = [e for e in etudiants if e[2] == "M"]
    filles = [e for e in etudiants if e[2] == "F"]
    
    # Gestion des cas limites
    total_garcons = len(garcons)
    total_etudiants = len(etudiants)
    nb_groupes = max(1, (total_etudiants + taille_groupe - 1) // taille_groupe)
    
    if min_garcons * nb_groupes > total_garcons:
        min_garcons = max(0, min(1, total_garcons))
        print(f"Ajustement automatique du nombre de garçons par groupe = {min_garcons} (seulement {total_garcons} garçons disponibles)")
    
    # Mélange et répartition initiale
    random.shuffle(garcons)
    random.shuffle(filles)
    
    while garcons or filles:
        groupe = []
        
        # Ajout des garçons d'abord
        for _ in range(min_garcons):
            if garcons:
                groupe.append(garcons.pop())
        
        # Compléter avec des filles
        while len(groupe) < taille_groupe and filles:
            groupe.append(filles.pop())
        
        # Si besoin, ajouter des garçons supplémentaires
        while len(groupe) < taille_groupe and garcons:
            groupe.append(garcons.pop())
        
        if groupe:  # Ne pas ajouter de groupe vide
            groupes.append(groupe)
    
    # Redistribution des restes si nécessaire
    if len(groupes) > 1:
        dernier_groupe = groupes[-1]
        
        # Vérifier si le dernier groupe est trop petit
        if len(dernier_groupe) < taille_groupe - 1:
            # Redistribuer les membres du dernier groupe
            for etudiant in dernier_groupe:
                redistribue = False
                
                # Essayer d'abord de mettre dans un groupe avec assez de garçons
                for g in groupes[:-1]:
                    garcons_dans_groupe = sum(1 for e in g if e[2] == "M")
                    if len(g) < taille_groupe and garcons_dans_groupe >= min_garcons:
                        g.append(etudiant)
                        redistribue = True
                        break
                
                # Sinon, trouver n'importe quel groupe avec de la place
                if not redistribue:
                    for g in groupes[:-1]:
                        if len(g) < taille_groupe:
                            g.append(etudiant)
                            redistribue = True
                            break
                
                # Si impossible à redistribuer, laisser dans le dernier groupe
                if not redistribue:
                    break
            
            # Si tous ont été redistribués, supprimer le dernier groupe
            if all(etudiant not in dernier_groupe for etudiant in dernier_groupe):
                groupes = groupes[:-1]
    
    # Équilibrer les groupes si possible
    i = 0
    while i < len(groupes):
        if len(groupes[i]) < taille_groupe - 1:
            for j in range(len(groupes)):
                if i != j and len(groupes[j]) > len(groupes[i]) + 1:
                    garcons_j = sum(1 for e in groupes[j] if e[2] == "M")
                    if garcons_j > min_garcons:
                        # Trouver un étudiant transférable
                        for k, etudiant in enumerate(groupes[j]):
                            if etudiant[2] == "F" or garcons_j > min_garcons:
                                groupes[i].append(groupes[j].pop(k))
                                break
        i += 1
    
    return groupes

def repartition_niveau(etudiants, taille_groupe, min_garcons):
    # Séparer les garçons et les filles triés par ordre alphabétique
    garcons = sorted([e for e in etudiants if e[2] == "M"], key=lambda x: x[0][0].upper())
    filles = sorted([e for e in etudiants if e[2] == "F"], key=lambda x: x[0][0].upper())
    
    # Vérifier si le nombre de garçons est suffisant
    nb_groupes = max(1, len(etudiants) // taille_groupe)
    if len(garcons) < min_garcons * nb_groupes:
        print(f"Attention: Impossible d'avoir {min_garcons} garçon(s) par groupe avec seulement {len(garcons)} garçons.")
        min_garcons = min(1, len(garcons))  # On réduit à 1 si possible
    
    groupes = []

    while garcons or filles:
        groupe = []
        initiales_groupe = set()
        
        # Ajouter d'abord les garçons
        for _ in range(min_garcons):
            if garcons:
                etudiant = garcons.pop(0)
                groupe.append(etudiant)
                initiales_groupe.add(etudiant[0][0].upper())
        
        # Ajouter des filles avec des initiales différentes
        while len(groupe) < taille_groupe and filles:
            for i, etudiant in enumerate(filles):
                if etudiant[0][0].upper() not in initiales_groupe:
                    groupe.append(filles.pop(i))
                    initiales_groupe.add(etudiant[0][0].upper())
                    break
            else:
                # Si toutes les filles restantes ont des initiales déjà présentes
                if filles:
                    groupe.append(filles.pop(0))
        
        # Si on manque de filles, ajouter des garçons supplémentaires
        while len(groupe) < taille_groupe and garcons:
            etudiant = garcons.pop(0)
            if etudiant[0][0].upper() not in initiales_groupe:
                groupe.append(etudiant)
                initiales_groupe.add(etudiant[0][0].upper())
        
        groupes.append(groupe)
    
    # Équilibrer les groupes
    i = 0
    while i < len(groupes):
        if len(groupes[i]) < taille_groupe - 1:
            for j in range(len(groupes)):
                if i != j and len(groupes[j]) > len(groupes[i]) + 1:
                    # Vérifier qu'on ne perturbe pas la répartition des garçons
                    if len([e for e in groupes[j] if e[2] == "M"]) > min_garcons:
                        # Trouver un étudiant transférable
                        for k, etudiant in enumerate(groupes[j]):
                            if etudiant[2] == "F" or len([e for e in groupes[j] if e[2] == "M"]) > min_garcons:
                                if etudiant[0][0].upper() not in {e[0][0].upper() for e in groupes[i]}:
                                    groupes[i].append(groupes[j].pop(k))
                                    break
        i += 1
    
    return groupes

def exporter_excel(grupos, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Répartition"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    group_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    
    ws.append(["Groupes", "Nom", "Prénom", "Genre"])
    for col in range(1, 5):
        ws.cell(row=1, column=col).font = header_font
        ws.cell(row=1, column=col).fill = header_fill
    
    row_num = 2
    for grupo_num, grupo in enumerate(grupos, 1):
        ws.cell(row=row_num, column=1, value=f"Groupe {grupo_num}")
        
        for estudiante in grupo:
            if isinstance(estudiante, tuple) and len(estudiante) == 3:
                nom, prenom, genre = estudiante
            elif isinstance(estudiante, str):
                nom = estudiante
                prenom = ""
                genre = ""
            else:
                nom = str(estudiante)
                prenom = ""
                genre = ""
            
            ws.cell(row=row_num, column=2, value=nom)
            ws.cell(row=row_num, column=3, value=prenom)
            ws.cell(row=row_num, column=4, value=genre)
            
            if grupo_num % 2 == 0:
                for col in range(1, 5):
                    ws.cell(row=row_num, column=col).fill = group_fill
            
            row_num += 1
        
        if len(grupo) > 1:
            ws.merge_cells(
                start_row=row_num - len(grupo), 
                end_row=row_num - 1, 
                start_column=1, 
                end_column=1
            )
    
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 10
    
    for merged_range in ws.merged_cells.ranges:
        for cell in merged_range.cells:
            ws.cell(row=cell[0], column=cell[1]).alignment = Alignment(
                horizontal='center', 
                vertical='center'
            )
    
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    wb.save(filepath)
    return filepath

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/choisir_entree', methods=['POST'])
def choisir_entree():
    choix = request.form.get('choix_entree')
    if choix == 'saisie':
        return redirect(url_for('saisie_manuelle'))
    elif choix == 'import':
        return redirect(url_for('importer_fichier'))
    else:
        flash('Veuillez choisir une option valide', 'error')
        return redirect(url_for('index'))

@app.route('/saisie_manuelle', methods=['GET', 'POST'])
def saisie_manuelle():
    global etudiants
    
    if request.method == 'POST':
        etudiants = []
        noms = request.form.getlist('nom[]')
        prenoms = request.form.getlist('prenom[]')
        genres = request.form.getlist('genre[]')
        
        for nom, prenom, genre in zip(noms, prenoms, genres):
            if nom.strip():  # Ne pas ajouter si le nom est vide
                etudiants.append((
                    nom.strip().capitalize(),
                    prenom.strip().capitalize(),
                    genre.upper()
                ))
        
        if len(etudiants) < 2:
            flash('Vous devez saisir au moins 2 étudiants', 'error')
            return render_template('saisie.html')
        
        return redirect(url_for('parametres_repartition'))
    
    return render_template('saisie.html')

@app.route('/importer_fichier', methods=['GET', 'POST'])
def importer_fichier():
    global etudiants
    
    if request.method == 'POST':
        if 'fichier' not in request.files:
            flash('Aucun fichier sélectionné', 'error')
            return redirect(request.url)
        
        fichier = request.files['fichier']
        if fichier.filename == '':
            flash('Aucun fichier sélectionné', 'error')
            return redirect(request.url)
        
        if fichier and allowed_file(fichier.filename):
            filename = secure_filename(fichier.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            fichier.save(filepath)
            
            try:
                wb = load_workbook(filepath)
                ws = wb.active
                etudiants = []
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    nom = str(row[0]).strip().title() if row and row[0] else ""
                    prenom = str(row[1]).strip().title() if len(row) > 1 and row[1] else ""
                    genre = str(row[2]).strip().upper()[0] if len(row) > 2 and row[2] else "M"
                    
                    if nom:
                        etudiants.append((nom, prenom, genre))
                
                if len(etudiants) < 2:
                    flash('Le fichier doit contenir au moins 2 étudiants valides', 'error')
                    return redirect(request.url)
                
                flash(f'{len(etudiants)} étudiants importés avec succès!', 'success')
                return redirect(url_for('parametres_repartition'))
            
            except Exception as e:
                flash(f'Erreur lors de la lecture du fichier: {str(e)}', 'error')
                return redirect(request.url)
        else:
            flash('Type de fichier non autorisé. Seuls les fichiers Excel sont acceptés.', 'error')
            return redirect(request.url)
    
    return render_template('import.html')

@app.route('/parametres_repartition', methods=['GET', 'POST'])
def parametres_repartition():
    global parametres
    
    if request.method == 'POST':
        taille_groupe = int(request.form.get('taille_groupe', 2))
        min_garcons = int(request.form.get('min_garcons', 1))
        type_repartition = request.form.get('type_repartition', 'aleatoire')
        
        if taille_groupe < 2:
            flash('La taille minimale d\'un groupe est 2', 'error')
            return redirect(request.url)
        
        parametres = {
            'taille_groupe': taille_groupe,
            'min_garcons': min_garcons,
            'type_repartition': type_repartition
        }
        
        return redirect(url_for('generer_groupes'))
    
    return render_template('parametres.html')


@app.route('/generer_groupes')
def generer_groupes():
    global etudiants, groupes, parametres

    if not parametres:
        flash('Veuillez d\'abord configurer les paramètres', 'error')
        return redirect(url_for('parametres_repartition'))
    
    if not etudiants or len(etudiants) < 2:
        flash('Pas assez d\'étudiants pour former des groupes', 'error')
        return redirect(url_for('index'))
    
    taille_groupe = parametres.get('taille_groupe', 2)
    min_garcons = parametres.get('min_garcons', 1)
    type_repartition = parametres.get('type_repartition', 'aleatoire')
    
    if type_repartition == 'aleatoire':
        groupes = repartition_aleatoire(etudiants, taille_groupe, min_garcons)
    else:
        groupes = repartition_niveau(etudiants, taille_groupe, min_garcons)
    
    # Ajout des paramètres au contexte du template
    return render_template('resultats.html', 
                         groupes=groupes,
                         taille_groupe=taille_groupe,
                         type_repartition=type_repartition.capitalize(),
                         min_garcons=min_garcons)

@app.route('/exporter', methods=['GET', 'POST'])
def exporter():
    global groupes
    
    if not groupes:
        flash("Aucun groupe à exporter. Veuillez d'abord générer des groupes.", 'error')
        return redirect(url_for('generer_groupes'))
    
    if request.method == 'POST':
        filename = request.form.get('filename', 'groupes.xlsx')
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        try:
            filepath = exporter_excel(groupes, filename)
            
            # Vérifier que le fichier a bien été créé
            if os.path.exists(filepath) and os.path.getsize(filepath) > 0:
                return send_from_directory(
                    app.config['UPLOAD_FOLDER'],
                    filename,
                    as_attachment=True,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            else:
                flash("Le fichier exporté est vide. Veuillez réessayer.", 'error')
                return redirect(url_for('exporter'))
                
        except Exception as e:
            flash(f"Erreur lors de l'export: {str(e)}", 'error')
            return redirect(url_for('exporter'))
    
    # Valeur par défaut pour le nom de fichier
    default_filename = f"groupes_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return render_template('export.html', default_filename=default_filename)

@app.route('/recommencer')
def recommencer():
    global etudiants, groupes, parametres
    etudiants = []
    groupes = []
    parametres = {}
    return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(debug=True)
